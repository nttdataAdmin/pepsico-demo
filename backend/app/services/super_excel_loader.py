"""Load all sheets from super_excel.xlsx into JSON-serializable structures."""
from __future__ import annotations

import os
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


def _cell_value(v: Any) -> Any:
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        try:
            return v.isoformat()
        except Exception:
            return str(v)
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v


def _sheet_to_records(ws) -> Dict[str, Any]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"columns": [], "rows": []}
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    # Deduplicate empty header names
    seen: Dict[str, int] = {}
    columns: List[str] = []
    for i, h in enumerate(header):
        name = h or f"column_{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        columns.append(name)
    out_rows: List[Dict[str, Any]] = []
    for raw in rows[1:]:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        row: Dict[str, Any] = {}
        for j, col in enumerate(columns):
            val = raw[j] if j < len(raw) else None
            row[col] = _cell_value(val)
        out_rows.append(row)
    return {"columns": columns, "rows": out_rows}


def default_excel_path() -> Path:
    env = os.getenv("SUPER_EXCEL_PATH", "").strip()
    if env:
        return Path(env)
    here = Path(__file__).resolve()
    for base in (here.parents[3], here.parents[2], here.parents[1]):
        candidate = base / "super_excel.xlsx"
        if candidate.exists():
            return candidate
    return here.parents[3] / "super_excel.xlsx"


def load_super_excel(path: Optional[Path] = None) -> Dict[str, Any]:
    p = path or default_excel_path()
    if not p.exists():
        return {"source": str(p), "sheets": {}, "error": "file_not_found"}

    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        sheets: Dict[str, Any] = {}
        for name in wb.sheetnames:
            ws = wb[name]
            rec = _sheet_to_records(ws)
            rec["row_count"] = len(rec["rows"])
            sheets[name] = rec
        return {"source": str(p.resolve()), "sheets": sheets, "sheet_names": list(sheets.keys())}
    finally:
        wb.close()


